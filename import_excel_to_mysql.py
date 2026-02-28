import datetime as dt
from pathlib import Path

import openpyxl
import pymysql


PROJECT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = PROJECT_DIR / "import_source.xlsx"

DB = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "hums",
    "charset": "utf8mb4",
    "autocommit": False,
}


def clean_text(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'"):
        s = s[1:]
    return s.strip()


def parse_date(value):
    s = clean_text(value)
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def map_status(value):
    s = clean_text(value)
    if not s:
        return "active"
    lower = s.lower()
    if "সক্রিয়" in s or "সক্রিয়" in s or "active" in lower:
        return "active"
    if "নিষ্ক্রিয়" in s or "নিষ্ক্রিয়" in s or "inactive" in lower:
        return "inactive"
    if "অপেক্ষ" in s or "pending" in lower:
        return "pending"
    return "active"


def to_int_or_none(value):
    s = clean_text(value)
    if not s:
        return None
    if s.isdigit():
        return int(s)
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
    except ValueError:
        pass
    return None


def main():
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, nid, program FROM beneficiaries")
    existing = {(str(nid), str(program)): int(i) for i, nid, program in cur.fetchall()}

    inserted = 0
    updated = 0
    skipped = 0

    insert_sql = """
        INSERT INTO beneficiaries (
            mis_number, name, name_en, gender, nid, program, union_name, phone, dob,
            father_en, father, mother_en, mother, spouse_name_en, spouse_name_bn,
            bank_mfs, account_number, age, division_name, district_name, upazila_name,
            ward_name, addr, status
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    update_sql = """
        UPDATE beneficiaries
        SET mis_number=%s, name=%s, name_en=%s, gender=%s, union_name=%s, phone=%s, dob=%s,
            father_en=%s, father=%s, mother_en=%s, mother=%s, spouse_name_en=%s, spouse_name_bn=%s,
            bank_mfs=%s, account_number=%s, age=%s, division_name=%s, district_name=%s, upazila_name=%s,
            ward_name=%s, addr=%s, status=%s
        WHERE id=%s
    """

    # Expected workbook columns (0-based):
    # 0=mis_number, 1=name_en, 2=name_bn, 3=gender, 4=dob, 5=father_en, 6=father_bn,
    # 7=mother_en, 8=mother_bn, 9=spouse_en, 10=spouse_bn, 11=program, 12=nid,
    # 13=status, 14=phone, 15=bank_mfs, 16=account_number, 17=age,
    # 18=division, 19=district, 20=upazila, 21=union, 22=ward, 23=address
    for row in ws.iter_rows(min_row=2, values_only=True):
        mis_number = clean_text(row[0] if len(row) > 0 else "")
        name_en = clean_text(row[1] if len(row) > 1 else "")
        name_bn = clean_text(row[2] if len(row) > 2 else "")
        gender = clean_text(row[3] if len(row) > 3 else "")
        dob = parse_date(row[4] if len(row) > 4 else "")
        father_en = clean_text(row[5] if len(row) > 5 else "")
        father_bn = clean_text(row[6] if len(row) > 6 else "")
        mother_en = clean_text(row[7] if len(row) > 7 else "")
        mother_bn = clean_text(row[8] if len(row) > 8 else "")
        spouse_en = clean_text(row[9] if len(row) > 9 else "")
        spouse_bn = clean_text(row[10] if len(row) > 10 else "")
        program = clean_text(row[11] if len(row) > 11 else "")
        nid = clean_text(row[12] if len(row) > 12 else "")
        status = map_status(row[13] if len(row) > 13 else "")
        phone = clean_text(row[14] if len(row) > 14 else "")
        bank_mfs = clean_text(row[15] if len(row) > 15 else "")
        account_number = clean_text(row[16] if len(row) > 16 else "")
        age = to_int_or_none(row[17] if len(row) > 17 else "")
        division_name = clean_text(row[18] if len(row) > 18 else "")
        district_name = clean_text(row[19] if len(row) > 19 else "")
        upazila_name = clean_text(row[20] if len(row) > 20 else "")
        union_name = clean_text(row[21] if len(row) > 21 else "")
        ward_name = clean_text(row[22] if len(row) > 22 else "")
        addr = clean_text(row[23] if len(row) > 23 else "")

        name = name_bn or name_en
        if not name or not program or not nid:
            skipped += 1
            continue

        key = (nid, program)
        if key in existing:
            cur.execute(
                update_sql,
                (
                    mis_number, name, name_en, gender, union_name, phone, dob,
                    father_en, father_bn, mother_en, mother_bn, spouse_en, spouse_bn,
                    bank_mfs, account_number, age, division_name, district_name,
                    upazila_name, ward_name, addr, status, existing[key],
                ),
            )
            updated += 1
        else:
            cur.execute(
                insert_sql,
                (
                    mis_number, name, name_en, gender, nid, program, union_name, phone, dob,
                    father_en, father_bn, mother_en, mother_bn, spouse_en, spouse_bn,
                    bank_mfs, account_number, age, division_name, district_name,
                    upazila_name, ward_name, addr, status,
                ),
            )
            inserted += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"Import complete. Inserted={inserted}, Updated={updated}, Skipped={skipped}")


if __name__ == "__main__":
    main()
